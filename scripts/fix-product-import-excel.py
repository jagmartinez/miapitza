"""Fix inventory import Excel: categories, SKUs, and units."""
import re
from openpyxl import load_workbook

SRC = r"c:\restaurant\Plantilla_Productos 20.05.26.xlsx"
OUT = r"c:\restaurant\Plantilla_Productos 20.05.26 - importacion.xlsx"

CAT_MAP = {
    "BEBIDAS": "Bebidas",
    "CONGELADOS": "Congelados",
    "EMPAQUES": "Empaques",
    "LIMPIEZA": "Limpieza",
    "MISCELÁNEOS": "Misceláneo",
    "MISCELANEOS": "Misceláneo",
    "VEGETALES": "Vegetales",
}

PREFIX = {
    "Bebidas": "BEB",
    "Congelados": "CON",
    "Empaques": "EMP",
    "Limpieza": "LIM",
    "Misceláneo": "MIS",
    "Vegetales": "VEG",
}

STORAGE_BY_CAT = {
    "Congelados": "FROZEN",
    "Vegetales": "PERISHABLE",
}

PERISHABLE_MISC = {
    "FRESAS",
    "HONGOS",
    "LECHE",
    "MANTECA",
    "PAN HAMBURGUESA",
    "PAN MOLIDO",
}


def normalize_key(value: str) -> str:
    return (
        value.strip()
        .upper()
        .replace("Á", "A")
        .replace("É", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ú", "U")
    )


def normalize_cat(raw: str) -> str:
    if not raw:
        return ""
    key = normalize_key(raw)
    for source, target in CAT_MAP.items():
        if normalize_key(source) == key:
            return target
    return raw.strip()


def infer_unit(name: str, cat_name: str) -> str:
    n = normalize_key(name)

    packaging_tokens = (
        "ENVASE", "VASO", "BOLSA", "PLATO", "CAJA", "PAPEL", "SERVILLETAS",
        "CUBIERTOS", "PAJILLAS", "EMPAQUE", "BARQUITO", "CERA", "ROLLO",
    )
    if cat_name == "Empaques" or any(token in n for token in packaging_tokens):
        if re.search(r"\b(PAQ\.?|PAQUETE|BOLSA|BOLSAS|SERVILLETAS)\b", n):
            return "paquete"
        if re.search(r"\bCAJA\b", n):
            return "caja"
        if re.search(r"\b(UNDS?|UNIDAD|DOCENA)\b", n):
            return "unidad"
        return "paquete"

    if re.search(r"\b(GALON|GAL\b|\d+\s*GL\b)", n):
        return "gal"
    if re.search(r"\b(LITROS?|LTS\b|\d+\s*LT\b|\d+\s*L\b)", n):
        return "l"
    if re.search(r"\b\d+\s*ML\b", n):
        return "ml"
    if re.search(r"\b(LBS?\b|LIBRA|LIB\b|POR LIBRA|X LIBRA|\d+\s*LB\b)", n):
        return "lb"
    if re.search(r"\b(OZ\b|ONZ\b|\d+\.?\d*\s*OZ\b)", n):
        return "oz"
    if re.search(r"\b\d+\.?\d*\s*G\b|\b\d+\s*GR\b", n):
        return "g"
    if re.search(r"\bKG\b|\bKILO\b", n):
        return "kg"

    if re.search(r"\bCAJA\b", n) and not re.search(r"\b\d+\s*(LB|LBS|KG|LT|LITRO)\b", n):
        return "caja"
    if re.search(r"\b(PAQ\.?|PAQUETE|BOLSA|BOLSAS|SERVILLETAS)\b", n):
        return "paquete"
    if re.search(r"\b(UNDS?|UNIDAD|DOCENA)\b", n):
        return "unidad"

    defaults = {
        "Bebidas": "unidad",
        "Vegetales": "lb",
        "Empaques": "paquete",
        "Limpieza": "unidad",
        "Congelados": "lb",
        "Misceláneo": "unidad",
    }
    return defaults.get(cat_name, "unidad")


def infer_storage(name: str, cat_name: str) -> str:
    if cat_name in STORAGE_BY_CAT:
        return STORAGE_BY_CAT[cat_name]
    if cat_name == "Vegetales":
        return "PERISHABLE"
    n = normalize_key(name)
    if cat_name == "Misceláneo":
        for token in PERISHABLE_MISC:
            if token in n:
                return "PERISHABLE"
    return "NON_PERISHABLE"


def main() -> None:
    wb = load_workbook(SRC)
    ws = wb["Productos"]
    counters = {prefix: 0 for prefix in PREFIX.values()}
    existing_skus: set[str] = set()

    if "Productos Actuales" in wb.sheetnames:
        existing = wb["Productos Actuales"]
        for row in range(2, existing.max_row + 1):
            value = existing.cell(row, 1).value
            if value:
                existing_skus.add(str(value).strip().upper())

    for row in range(2, ws.max_row + 1):
        name = (ws.cell(row, 2).value or "").__str__().strip()
        if not name:
            continue

        raw_cat = (ws.cell(row, 3).value or "").__str__().strip()
        cat = normalize_cat(raw_cat)
        unit = infer_unit(name, cat)
        storage = infer_storage(name, cat)

        prefix = PREFIX.get(cat, "ING")
        counters[prefix] = counters.get(prefix, 0) + 1
        sku = f"{prefix}-{counters[prefix]:06d}"
        while sku.upper() in existing_skus:
            counters[prefix] += 1
            sku = f"{prefix}-{counters[prefix]:06d}"
        existing_skus.add(sku.upper())

        ws.cell(row, 1).value = sku
        ws.cell(row, 3).value = cat
        ws.cell(row, 4).value = unit
        if not (ws.cell(row, 8).value or "").__str__().strip():
            ws.cell(row, 8).value = "INGREDIENT"
        ws.cell(row, 9).value = storage

    categories_sheet = next(s for s in wb.sheetnames if s.lower().startswith("categor"))
    cs = wb[categories_sheet]
    cs.delete_rows(2, cs.max_row)
    cs.cell(1, 3).value = "Prefijo SKU"
    category_rows = [
        ("Bebidas", "Bebidas y líquidos", "BEB"),
        ("Congelados", "Productos congelados y refrigerados", "CON"),
        ("Empaques", "Empaques, envases y desechables", "EMP"),
        ("Limpieza", "Productos de limpieza y aseo", "LIM"),
        ("Misceláneo", "Productos varios y misceláneos", "MIS"),
        ("Vegetales", "Vegetales, verduras y hortalizas", "VEG"),
        ("Carnes", "Carnes rojas, blancas y embutidos", "CAR"),
        ("Lácteos", "Productos lácteos y derivados", "LAC"),
        ("Entradas", "Appetizers", ""),
        ("Platos Fuertes", "Main Courses", ""),
        ("Postres", "Desserts", ""),
    ]
    for index, (name, description, code_prefix) in enumerate(category_rows, start=2):
        cs.cell(index, 1).value = name
        cs.cell(index, 2).value = description
        cs.cell(index, 3).value = code_prefix

    wb.save(OUT)

    valid_cats = set(CAT_MAP.values()) | {"Carnes", "Lácteos", "Entradas", "Platos Fuertes", "Postres"}
    valid_units = {
        "g", "kg", "mg", "lb", "oz", "qq", "arr", "ml", "l", "oz_fl", "gal",
        "unidad", "paquete", "caja", "saco", "docena",
    }
    errors = 0
    total = 0
    for row in range(2, ws.max_row + 1):
        name = (ws.cell(row, 2).value or "").__str__().strip()
        if not name:
            continue
        total += 1
        sku = (ws.cell(row, 1).value or "").__str__().strip()
        cat = (ws.cell(row, 3).value or "").__str__().strip()
        unit = (ws.cell(row, 4).value or "").__str__().strip()
        if not sku or not unit or cat not in valid_cats or unit not in valid_units:
            print("ERR", row, sku, name, cat, unit)
            errors += 1

    print(f"saved {OUT}")
    print(f"rows {total}, errors {errors}")
    print("sku counts", {key: value for key, value in counters.items() if value})


if __name__ == "__main__":
    main()
